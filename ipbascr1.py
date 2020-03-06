import time
import os
import shutil
from selenium import webdriver
import os
from datetime import datetime
datestring = datetime.strftime(datetime.now(), '(%Y-%m-%d)-(%H.%M.ss)')
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import autopy


cwd = os.getcwd()
DRIVER = 'chromedriver'

pathmanual = "D:/IPBA 1/"

chrome_options = webdriver.ChromeOptions()
if os.name == "nt":
    # If current OS is Windows
    chrome_options.add_argument("--start-maximized")
else:
    # Other OS (Mac / Linux)
    chrome_options.add_argument("--kiosk")

driver = webdriver.Chrome(DRIVER, chrome_options = chrome_options)

while True:
	try:
		driver.get('http://www.ibpa.co.id/DataPasarSuratUtang/Indeks/INDOBeX.aspx')
		time.sleep(5)

		all_html = driver.page_source
		soup = BeautifulSoup(all_html,"html.parser")

		datasets = []

		date = soup.find("span",{"id": "dnn_ctr695_INDOBEX_Data_lblDate"},text = True)
		day = soup.find("span",{"id": "dnn_ctr695_INDOBEX_Data_lblDay"},text = True)

		all_date = day.get_text()+" "+date.get_text() + " \n"
		print(all_date)
		break

	except:
		print('try again')
		






#import datetime---------------------------------------------------------------------------------
server_time_date = date.get_text()[:2]
server_time_month = date.get_text()[3:6]
server_time_year = date.get_text()[9:] 
server_time = server_time_date + " " + server_time_month + " " + server_time_year
print(server_time)

import datetime
date_xls = datetime.datetime.strptime(server_time, '%d %b %Y')
print(date_xls)
#import datetime---------------------------------------------------------------------------------






index_x = -1
writing_xlsx = 0
wb = load_workbook('D:\\Project Kantor\\IPBA_UPLOAD\\original.xlsx')
ws = wb['Sheet 1']
for table in soup.findAll("table", {"id": "dnn_ctr695_INDOBEX_Data_gvDailyDate"}):
	for row in table.findAll("tr")[3:]:
			
			isinya = [td.get_text().replace("  ","") for td in row.findAll("td") if td.get_text()]
			del isinya[:3]
			del isinya[1:]
			p_isinya = ''.join(map(str, isinya))
			p_isinya = p_isinya.replace('.','.')
			datasets.append(p_isinya)

			alpha = 'D'
			tanggal = 'C'
			mix = alpha+str(index_x)
			mix_tanggal = tanggal+str(index_x)

			if isinya == []:
				datasets.remove(p_isinya)
			if index_x >= 2 and index_x != 17 and writing_xlsx >= 3:
				print(mix)
				print(p_isinya)
				ws[mix].value = float(p_isinya)
				sellll = ws.cell(column=4, row=index_x)
				sellll.number_format = 'General'
				ws[mix_tanggal].value = date_xls
				
			
			index_x += 1
			writing_xlsx += 1
print('------------------------------------')
print(*datasets, sep = "\n")


#--------------------------------------------------------#--------------------------------------------------------#--------------------------------------------------------

while True:
	try:

		driver.get('http://www.ibpa.co.id/DataPasarSuratUtang/Indeks/IndonesiaSukukIndex/tabid/202/Default.aspx')
		time.sleep(10)

		all_html_sukuk = driver.page_source
		soup_sukuk = BeautifulSoup(all_html_sukuk,"html.parser")

		datasets_sukuk = []

		date = soup_sukuk.find("span",{"id": "dnn_ctr713_INDEX_Sukuk_Data_lblDate"},text = True)
		day = soup_sukuk.find("span",{"id": "dnn_ctr713_INDEX_Sukuk_Data_lblDay"},text = True)

		all_date = day.get_text()+" "+date.get_text() + " \n"
		print(all_date)
		break

	except:
		print('try again')


#----------------------------------------------------------------------------------------------------------------------
server_time_date = date.get_text()[:2]
server_time_month = date.get_text()[3:6]
server_time_year = date.get_text()[9:] 
server_time = server_time_date + " " + server_time_month + " " + server_time_year
print(server_time)


import datetime
date_xls = datetime.datetime.strptime(server_time, '%d %b %Y')
print(date_xls)

#----------------------------------------------------------------------------------------------------------------------



index_x_sukuk = 14
writing_xlsx_sukuk = 16
for table_sukuk in soup_sukuk.findAll("table", {"id": "dnn_ctr713_INDEX_Sukuk_Data_gvDailyDate"}):
	for row_sukuk in table_sukuk.findAll("tr")[3:]:
		isinya_sukuk = [td.get_text().replace("  ","") for td in row_sukuk.findAll("td") if td.get_text()]
		del isinya_sukuk[:3]
		del isinya_sukuk[1:]
		p_isinya_sukuk = ''.join(map(str, isinya_sukuk))
		p_isinya_sukuk = p_isinya_sukuk.replace('.','.')
		datasets_sukuk.append(p_isinya_sukuk)

		alpha = 'D'
		tanggal = 'C'
		mix = alpha+str(index_x_sukuk)
		mix_tanggal_sukuk = tanggal+str(index_x_sukuk)
		if isinya_sukuk == []:
			datasets_sukuk.remove(p_isinya_sukuk)
		if index_x_sukuk >= 17  and index_x_sukuk != 33:
			print(mix)
			print(p_isinya_sukuk)
			ws[mix].value = float(p_isinya_sukuk)
			sellll = ws.cell(column=4, row=index_x_sukuk)
			sellll.number_format = 'General'
			ws[mix_tanggal_sukuk].value = date_xls

			
			# ws[mix] = 
		
		index_x_sukuk += 1
		writing_xlsx_sukuk += 1

print('------------------------------------')


from datetime import datetime
judul = datetime.strftime(date_xls, '%Y_%m_%d')

wb.save('D:\\Project Kantor\\IPBA_UPLOAD\\By_Code_macro-uploader'+judul+'_INDOBEX.xlsx')		

print(*datasets_sukuk, sep = "\n")


while True:
	try:
		print('doing step 11, upload to website')

		#login dulu aja
		driver.get('https://pasardana.id/admin/')     
		for n in range(82):
			time.sleep(0.25)
			print('time.sleep(0.25)')
		username = driver.find_element_by_name('username')
		username.send_keys("admin_upload")
		for q in range(10):
			time.sleep(0.25)
			print('time.sleep(0.25)')
		password = driver.find_element_by_name('password')
		password.send_keys("alfamart312")
		for r in range(8):
			time.sleep(0.25)
			print('time.sleep(0.25)')

		autopy.mouse.smooth_move(0,0)
		autopy.mouse.click()
		autopy.mouse.smooth_move(1200,440)
		autopy.mouse.click()

		print('button login has been clicked')



		#Upload to website
		
		for g in range(14):
			time.sleep(0.25)
			print('time.sleep(0.25)')


		driver.get('https://pasardana.id/admin/macro/data')		
		for h in range(120):
			time.sleep(0.25)
			print('time.sleep(0.25)')
		wait = WebDriverWait(driver, 10)
		wait.until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Upload"]'))).click()
		for i in range(14):
			time.sleep(0.25)
			print('time.sleep(0.25)')
		driver.find_element_by_css_selector("input[ng-model='files']").send_keys("D:\\Project Kantor\\IPBA_UPLOAD\\By_Code_macro-uploader"+judul+"_INDOBEX.xlsx") 

		print('file has been imported')

		#12. klik button save
		# autopy.mouse.smooth_move(0,0)
		autopy.mouse.smooth_move(0,0)
		autopy.mouse.click()
		autopy.mouse.smooth_move(595,495)
		autopy.mouse.click()

		print('button save has been clicked')
		break
	except:
		driver.close()
		print('try again open')
		# session.close()
		for m in range(22):
			time.sleep(0.25)
			print('time.sleep(0.25)')

    
for b in range(6):
	time.sleep(0.25)
	print('time.sleep(0.25)')

driver.close()
